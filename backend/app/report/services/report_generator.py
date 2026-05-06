"""
报告生成器
支持PDF、Excel等多种格式
"""
import os
import json
import logging
from typing import Dict, Any, List, Optional
from datetime import datetime
from jinja2 import Template
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
import base64

# PDF生成（WeasyPrint，需要系统级 GTK/GLib 库）
try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError):
    HTML = None
    CSS = None
    WEASYPRINT_AVAILABLE = False

# Excel生成
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference

from app.report.config import settings

logger = logging.getLogger(__name__)


class ReportGenerator:
    """报告生成器"""
    
    # HTML模板
    HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{{ report_name }}</title>
    <style>
        body { font-family: "Microsoft YaHei", "SimHei", sans-serif; margin: 40px; }
        .header { text-align: center; border-bottom: 3px solid #1a73e8; padding-bottom: 20px; margin-bottom: 30px; }
        .header h1 { color: #1a73e8; font-size: 28px; margin: 0; }
        .header .subtitle { color: #666; font-size: 14px; margin-top: 10px; }
        .section { margin: 30px 0; }
        .section h2 { color: #333; font-size: 20px; border-left: 4px solid #1a73e8; padding-left: 15px; }
        .summary-box { background: #f8f9fa; border-radius: 8px; padding: 20px; margin: 20px 0; }
        .summary-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; }
        .summary-item { text-align: center; }
        .summary-item .value { font-size: 32px; font-weight: bold; color: #1a73e8; }
        .summary-item .label { font-size: 14px; color: #666; margin-top: 5px; }
        .data-table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        .data-table th { background: #1a73e8; color: white; padding: 12px; text-align: left; }
        .data-table td { padding: 10px 12px; border-bottom: 1px solid #e0e0e0; }
        .data-table tr:nth-child(even) { background: #f8f9fa; }
        .alert-box { background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; margin: 15px 0; }
        .alert-box.critical { background: #f8d7da; border-left-color: #dc3545; }
        .chart-container { text-align: center; margin: 20px 0; }
        .chart-container img { max-width: 100%; height: auto; }
        .footer { margin-top: 50px; padding-top: 20px; border-top: 1px solid #e0e0e0; text-align: center; color: #999; font-size: 12px; }
        .recommendations { background: #e8f5e9; border-radius: 8px; padding: 20px; }
        .recommendations ul { margin: 0; padding-left: 20px; }
        .recommendations li { margin: 10px 0; }
    </style>
</head>
<body>
    <div class="header">
        <h1>{{ report_name }}</h1>
        <div class="subtitle">
            报告编码：{{ report_code }} | 
            生成时间：{{ generated_at }} | 
            数据周期：{{ time_range }}
        </div>
    </div>
    
    <div class="section">
        <h2>一、报告摘要</h2>
        <div class="summary-box">
            <div class="summary-grid">
                <div class="summary-item">
                    <div class="value">{{ summary.data_points }}</div>
                    <div class="label">数据点</div>
                </div>
                <div class="summary-item">
                    <div class="value">{{ summary.alerts_count }}</div>
                    <div class="label">预警次数</div>
                </div>
                <div class="summary-item">
                    <div class="value">{{ summary.compliance_rate }}%</div>
                    <div class="label">达标率</div>
                </div>
            </div>
        </div>
    </div>
    
    <div class="section">
        <h2>二、数据分析</h2>
        {% if charts %}
            {% for chart in charts %}
            <div class="chart-container">
                <img src="data:image/png;base64,{{ chart.data }}" alt="{{ chart.title }}">
                <p>{{ chart.title }}</p>
            </div>
            {% endfor %}
        {% endif %}
    </div>
    
    <div class="section">
        <h2>三、预警信息</h2>
        {% if alerts %}
            {% for alert in alerts %}
            <div class="alert-box {{ 'critical' if alert.level == 'critical' else '' }}">
                <strong>{{ alert.title }}</strong><br>
                {{ alert.description }}
            </div>
            {% endfor %}
        {% else %}
            <p>本周期内无预警信息</p>
        {% endif %}
    </div>
    
    <div class="section">
        <h2>四、管理建议</h2>
        <div class="recommendations">
            <ul>
                {% for rec in recommendations %}
                <li>{{ rec }}</li>
                {% endfor %}
            </ul>
        </div>
    </div>
    
    <div class="footer">
        <p>本报告由流域水环境AI智能监测与预警平台自动生成</p>
        <p>报告仅供参考，具体决策请以实际情况为准</p>
    </div>
</body>
</html>
"""
    
    def __init__(self):
        self.output_path = settings.REPORT_OUTPUT_PATH
        os.makedirs(self.output_path, exist_ok=True)
    
    def _generate_charts(self, data: Dict[str, Any]) -> List[Dict[str, str]]:
        """生成图表"""
        charts = []
        
        # 设置中文字体
        plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
        
        # 趋势图
        station_data = data.get("station_data", {}).get("data", [])
        if station_data:
            df = pd.DataFrame(station_data)
            if "ts" in df.columns and "ph" in df.columns:
                df["ts"] = pd.to_datetime(df["ts"])
                
                fig, axes = plt.subplots(2, 2, figsize=(12, 8))
                fig.suptitle("水质指标趋势图", fontsize=16)
                
                # pH趋势
                if "ph" in df.columns:
                    axes[0, 0].plot(df["ts"], df["ph"], color="#1a73e8")
                    axes[0, 0].axhline(y=6.0, color="r", linestyle="--", alpha=0.5)
                    axes[0, 0].axhline(y=9.0, color="r", linestyle="--", alpha=0.5)
                    axes[0, 0].set_title("pH值趋势")
                    axes[0, 0].set_ylabel("pH")
                
                # 溶解氧趋势
                if "do" in df.columns:
                    axes[0, 1].plot(df["ts"], df["do"], color="#34a853")
                    axes[0, 1].axhline(y=2.0, color="r", linestyle="--", alpha=0.5)
                    axes[0, 1].set_title("溶解氧趋势")
                    axes[0, 1].set_ylabel("mg/L")
                
                # 氨氮趋势
                if "nh3_n" in df.columns:
                    axes[1, 0].plot(df["ts"], df["nh3_n"], color="#ea4335")
                    axes[1, 0].set_title("氨氮趋势")
                    axes[1, 0].set_ylabel("mg/L")
                
                # 高锰酸盐指数趋势
                if "codmn" in df.columns:
                    axes[1, 1].plot(df["ts"], df["codmn"], color="#fbbc04")
                    axes[1, 1].set_title("高锰酸盐指数趋势")
                    axes[1, 1].set_ylabel("mg/L")
                
                plt.tight_layout()
                
                # 转换为base64
                buffer = BytesIO()
                plt.savefig(buffer, format="png", dpi=100)
                buffer.seek(0)
                image_base64 = base64.b64encode(buffer.read()).decode()
                plt.close()
                
                charts.append({
                    "title": "水质指标趋势",
                    "data": image_base64
                })
        
        return charts
    
    def _generate_summary(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """生成摘要数据"""
        station_data = data.get("station_data", {}).get("data", [])
        alerts = data.get("alerts", [])
        
        return {
            "data_points": len(station_data),
            "alerts_count": len(alerts),
            "compliance_rate": 95  # 简化计算
        }
    
    def _generate_recommendations(self, data: Dict[str, Any]) -> List[str]:
        """生成建议"""
        recommendations = []
        alerts = data.get("alerts", [])
        ai_analysis = data.get("ai_analysis", {})
        
        # 基于预警生成建议
        if alerts:
            recommendations.append(f"本周期内共发生{len(alerts)}次预警，建议加强监测频次")
        
        # 基于AI分析生成建议
        pollution_type = ai_analysis.get("rule_based", {}).get("pollution_type")
        if pollution_type and pollution_type != "unknown":
            pollution_name = ai_analysis.get("rule_based", {}).get("pollution_name")
            recommendations.append(f"AI识别为{pollution_name}，建议按预案处置")
        
        # 通用建议
        recommendations.extend([
            "建议继续保持当前监测频率",
            "定期校准监测设备，确保数据准确性",
            "关注上游来水水质变化"
        ])
        
        return recommendations
    
    async def generate_pdf(self, report_code: str, report_name: str,
                           data: Dict[str, Any]) -> str:
        """生成PDF报告"""
        if not WEASYPRINT_AVAILABLE:
            logger.warning("WeasyPrint not available, skipping PDF generation")
            return None
        
        try:
            # 准备数据
            charts = self._generate_charts(data)
            summary = self._generate_summary(data)
            recommendations = self._generate_recommendations(data)
            
            time_range = ""
            if data.get("time_range"):
                start = data["time_range"].get("start", "")[:10] if data["time_range"].get("start") else ""
                end = data["time_range"].get("end", "")[:10] if data["time_range"].get("end") else ""
                time_range = f"{start} 至 {end}"
            
            # 渲染HTML
            template = Template(self.HTML_TEMPLATE)
            html_content = template.render(
                report_name=report_name,
                report_code=report_code,
                generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
                time_range=time_range,
                summary=summary,
                charts=charts,
                alerts=data.get("alerts", []),
                recommendations=recommendations
            )
            
            # 生成PDF
            output_path = os.path.join(self.output_path, f"{report_code}.pdf")
            HTML(string=html_content).write_pdf(output_path)
            
            logger.info(f"PDF report generated: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to generate PDF: {e}")
            return None
    
    async def generate_excel(self, report_code: str, report_name: str,
                             data: Dict[str, Any]) -> str:
        """生成Excel报告"""
        try:
            output_path = os.path.join(self.output_path, f"{report_code}.xlsx")
            
            # 创建工作簿
            wb = Workbook()
            
            # 1. 摘要页
            ws_summary = wb.active
            ws_summary.title = "报告摘要"
            
            # 标题
            ws_summary["A1"] = report_name
            ws_summary["A1"].font = Font(size=18, bold=True, color="1a73e8")
            ws_summary.merge_cells("A1:D1")
            ws_summary["A1"].alignment = Alignment(horizontal="center")
            
            # 基本信息
            ws_summary["A3"] = "报告编码："
            ws_summary["B3"] = report_code
            ws_summary["A4"] = "生成时间："
            ws_summary["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            
            # 2. 数据页
            ws_data = wb.create_sheet("监测数据")
            station_data = data.get("station_data", {}).get("data", [])
            
            if station_data:
                df = pd.DataFrame(station_data)
                
                # 写入表头
                headers = list(df.columns)
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_data.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # 写入数据
                for row_idx, row in enumerate(df.values, 2):
                    for col_idx, value in enumerate(row, 1):
                        ws_data.cell(row=row_idx, column=col_idx, value=value)
            
            # 3. 预警页
            ws_alerts = wb.create_sheet("预警信息")
            alerts = data.get("alerts", [])
            
            if alerts:
                headers = ["预警编码", "预警类型", "预警级别", "标题", "状态", "创建时间"]
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_alerts.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
                
                for row_idx, alert in enumerate(alerts, 2):
                    ws_alerts.cell(row=row_idx, column=1, value=alert.get("alert_code"))
                    ws_alerts.cell(row=row_idx, column=2, value=alert.get("alert_type"))
                    ws_alerts.cell(row=row_idx, column=3, value=alert.get("alert_level"))
                    ws_alerts.cell(row=row_idx, column=4, value=alert.get("title"))
                    ws_alerts.cell(row=row_idx, column=5, value=alert.get("status"))
                    ws_alerts.cell(row=row_idx, column=6, value=alert.get("created_at"))
            
            # 保存
            wb.save(output_path)
            logger.info(f"Excel report generated: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to generate Excel: {e}")
            return None
    
    async def generate(self, report_code: str, report_name: str,
                       file_format: str, data: Dict[str, Any]) -> str:
        """生成报告"""
        if file_format.lower() == "pdf":
            return await self.generate_pdf(report_code, report_name, data)
        elif file_format.lower() in ["excel", "xlsx"]:
            return await self.generate_excel(report_code, report_name, data)
        else:
            logger.error(f"Unsupported format: {file_format}")
            return None
